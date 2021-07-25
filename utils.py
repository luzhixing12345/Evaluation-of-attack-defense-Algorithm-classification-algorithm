from numpy import show_config
import openpyxl
from openpyxl.utils import get_column_letter
import os
#关于python数据写入excel
#https://www.bilibili.com/video/BV1m4411K7Tc?from=search&seid=8211552785915364465

excel_name = 'result.xlsx'

def change_data_size(cfg,data):
    new_data = data.view(cfg.batch_size_test,cfg.height,cfg.width,cfg.channel)
    return new_data

def rechange_data_size(cfg,data):
    new_data = data.view(cfg.batch_size_test,cfg.channel,cfg.height,cfg.width,)
    return new_data


def check(save_pos):
    return os.path.exists(save_pos)

def remove_Sheet(wb,sheet_list):
    if 'Sheet' in sheet_list:
        wb.remove(wb['Sheet'])
    if 'Sheet1' in sheet_list:
        wb.remove(wb['Sheet1'])
def sheet_init(sheet):
    for row in range (3,7):
        for col in range (2,7):
            sheet.column_dimensions[get_column_letter(col)].width = 20.0
            sheet.row_dimensions[row].height = 40
    sheet['C3']='PURE_VALUE'
    sheet['D3']='AFTER_ATTACK'
    sheet['E3']='BDR'
    sheet['F3']='JPEG'
    sheet['B4']='MIM'
    sheet['B5']='FGSM'
    sheet['B6']='DEEPFOOL'

def save_data(cfg,pure_result,adv_result,def_result,sheet):
    from attack import attack_set
    from defense import defense_set

    #数据存储形式为B3-F6 4x5的一个excel矩阵
    row = str(4 + list(attack_set.keys()).index(cfg.attack_method))
    column = chr(ord('E') + list(defense_set.keys()).index(cfg.defense_method))
    sheet['C'+row]=pure_result
    sheet['D'+row]=adv_result
    sheet[column+row]=def_result

def save_config(cfg,pure_result,adv_result,def_result,sheet):

    start_row = str(sheet.max_row+2)
    sheet['B'+start_row]='PURE'
    sheet['C'+start_row]=cfg.attack_method
    sheet['D'+start_row]=cfg.defense_method

    start_row = str(int(start_row)+1)
    sheet['B'+start_row]=pure_result
    sheet['C'+start_row]=adv_result
    sheet['D'+start_row]=def_result

    start_row =int(start_row)+2
    sheet['B'+str(start_row)]='ATTACK_ARGUMENTS'
    for col in range (len(cfg.attack_arguments)):
        sheet.cell(column= 3+2*col,row=start_row,value=list(cfg.attack_arguments.keys())[col])
        sheet.cell(column= 4+2*col,row=start_row,value=list(cfg.attack_arguments.values())[col])

    start_row =int(start_row)+2
    sheet['B'+str(start_row)]='DEFENSE_ARGUMENTS'
    for col in range (len(cfg.defense_arguments)):
        sheet.cell(column= 3+2*col,row=start_row,value=list(cfg.defense_arguments.keys())[col])
        sheet.cell(column= 4+2*col,row=start_row,value=list(cfg.defense_arguments.values())[col])

def record(cfg,pure_result,adv_result,def_result,save_pos= excel_name):
    if check(save_pos):
        wb = openpyxl.load_workbook(save_pos)
    else :
        wb = openpyxl.Workbook()
    sheet_list = wb.sheetnames
    sheet_name = cfg.network + cfg.dataset 
    if sheet_name in sheet_list:
        sheet = wb[sheet_name]
    else :
        sheet = wb.create_sheet(title = sheet_name)
        sheet_init(sheet)

    save_data(cfg,pure_result,adv_result,def_result,sheet)
    save_config(cfg,pure_result,adv_result,def_result,sheet)
    remove_Sheet(wb,sheet_list)
    wb.save(save_pos)